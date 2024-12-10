"""
# 셀레니움_기본설정
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

# 크롬 드라이버 자동 업데이트
from webdriver_manager.chrome import ChromeDriverManager

# 브라우저 꺼짐 방지
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

# 불필요한 에러 메시지 없애기
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

# 웹페이지 주소 이동
driver.get("http://www.naver.com")

# 3. 네이버 로그인 자동화

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

# 크롬 드라이버 자동 업데이트
from webdriver_manager.chrome import ChromeDriverManager

import time
import pyautogui
import pyperclip

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

# 불필요한 에러 메시지 없애기
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

# 웹페이지 주소 이동
driver.implicitly_wait(5)  # 웹페이지가 로딩 될때까지 5초는 기다림
driver.maximize_window()  # 화면 최대화
driver.get("https://nid.naver.com/nidlogin.login?mode=form&url=https://www.naver.com/")

# 아이디 입력창
id = driver.find_element(By.CSS_SELECTOR, "#id")
id.click()
# id.send_keys("tldus0_0")
pyperclip.copy("tldus0_0")
pyautogui.hotkey("ctrl", "v")
time.sleep(2)

# 비밀번호 입력창
pw = driver.find_element(By.CSS_SELECTOR, "#pw")
pw.click()
# pw.send_keys("   ")
pyperclip.copy("   ")
pyautogui.hotkey("ctrl", "v")
time.sleep(2)

# 로그인 버튼
login_btn = driver.find_element(By.CSS_SELECTOR, "#log\.login")
login_btn.click()
""" """
import requests
from bs4 import BeautifulSoup

# naver 서버에 대화를 시도
response = requests.get("https://finance.naver.com/")

# naver에서 html 줌
html = response.text

soup = BeautifulSoup(html, "html.parser")

# id 값이 NM_set_home_btn인 놈 한개를 찾아냄
Word = soup.select_one("#menu > ul > li.m1.first.on > a > span.tx")

# 텍스트 요소만 출력
print(Word)
"""
"""
# 뉴스제목
import requests
from bs4 import BeautifulSoup

response = requests.get(
    "https://search.naver.com/search.naver?where=news&sm=tab_jum&query=%EC%82%BC%EC%84%B1%EC%A0%84%EC%9E%90"
)
html = response.text
soup = BeautifulSoup(html, "html.parser")
links = soup.select(".news_tit")  # 결과는 리스트

for link in links:
    title = link.text  # 태그 안에 텍스트요소를 가져온다
    url = link.attrs["href"]  # href의 속성값을 가져온다
    print(title, url)
"""
"""
# 검색어 변경하기 01
import requests
from bs4 import BeautifulSoup

keyword = input("검색어를 입력하세요>>>")
response = requests.get(
    "https://search.naver.com/search.naver?where=news&sm=tab_jum&query=" + keyword
)
html = response.text
soup = BeautifulSoup(html, "html.parser")
links = soup.select(".news_tit")  # 결과는 리스트

for link in links:
    title = link.text  # 태그 안에 텍스트요소를 가져온다
    url = link.attrs["href"]  # href의 속성값을 가져온다
    print(title, url)
"""
"""
# 검색어 변경하기 02
import requests
from bs4 import BeautifulSoup
import pyautogui

keyword = pyautogui.prompt("검색어를 입력하세요.")
response = requests.get(
    f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={keyword}"
)
html = response.text
soup = BeautifulSoup(html, "html.parser")
links = soup.select(".news_tit")  # 결과는 리스트

for link in links:
    title = link.text  # 태그 안에 텍스트요소를 가져온다
    url = link.attrs["href"]  # href의 속성값을 가져온다
    print(title, url)
"""
"""
# 여러페이지 가저오기
import requests
from bs4 import BeautifulSoup
import pyautogui

keyword = pyautogui.prompt("검색어를 입력하세요.")
lastpage = pyautogui.prompt("마지막 페이지번호를 입력해주세요")
pageNum = 1
for i in range(1, int(lastpage) * 10, 10):
    print(f"\n{pageNum}페이지입니다.\n")
    response = requests.get(
        f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={keyword}&start={i}"
    )
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    links = soup.select(".news_tit")  # 결과는 리스트

    for link in links:
        title = link.text  # 태그 안에 텍스트요소를 가져온다
        url = link.attrs["href"]  # href의 속성값을 가져온다
        print(title, url)
    pageNum = pageNum + 1
"""
"""
# 1, 데이터 추출
import requests
from bs4 import BeautifulSoup

# 종목 리스트
codes = ["005930", "000660", "035720"]

for code in codes:
    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    price = soup.select_one("#_nowVal").text
    price = price.replace(",", "")
    print(price)

"""
"""
import requests
from bs4 import BeautifulSoup
import openpyxl

fpath = r"C:\git_files\Study\data.xlsx"

wb = openpyxl.Workbook()
codes = ["005930", "000660", "035720"]
ws = wb.create_sheet("stock")

i = 2
for code in codes:
    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    name = soup.select_one("#middle > div.h_company > div.wrap_company > h2 > a").text
    ws[f"A{i}"] = name
    price = soup.select_one("#_nowVal").text
    price = price.replace(",", "")
    ws[f"B{i}"] = price
    i += 1

wb.save(fpath)
"""
"""
# 파이썬 엑셀 다루기
import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()
# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet("오징어게임")
# 3) 데이터 추가하기
ws["A1"] = "참가번호"
ws["B1"] = "성명"

ws["A2"] = 1
ws["B2"] = "오일남"

# 4) 엑셀 저장하기
wb.save("C:\git_files\Study\data.xlsx")
""" """
import openpyxl

fpath = r"C:\git_files\Study\data.xlsx"

# 1) 엑셀 만들기
wb = openpyxl.load_workbook(fpath)

# 2) 엑셀 시트 선택
ws = wb["오징어게임"]

# 3) 데이터 수정하기
ws["A3"] = 456
ws["B3"] = "성기훈"

# 4) 엑셀 저장하기
wb.save(fpath)
""" """
# 쿠팡 웹 크롤링 (광고상품 제외)
import requests
from bs4 import BeautifulSoup

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Accept-Language": "ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3",
}
main_url = "https://www.coupang.com/np/search?rocketAll=false&searchId=d0f2d44aef82434ba6bf5356b6d9875c&q=%EA%B2%8C%EC%9D%B4%EB%B0%8D%EB%A7%88%EC%9A%B0%EC%8A%A4&brand=&offerCondition=&filter=&availableDeliveryFilter=&filterType=&isPriceRange=false&priceRange=&minPrice=&maxPrice=&page=1&trcid=&traid=&filterSetByUser=true&channel=user&backgroundColor=&searchProductCount=131227&component=&rating=0&sorter=scoreDesc&listSize=36"

# 헤더에 User-Agent를 추가하지 않으면 오류가 남(멈춰버림)
response = requests.get(main_url, headers=headers)
html = response.text
soup = BeautifulSoup(html, "html.parser")
links = soup.select("a.search-product-link")  # select의 결과는 리스트 자료형
print(links)
for link in links:
    # 광고상품 제거
    if len (link.select("span.ad-badge-text")) > 0:
        print("광고상품입니다.")
    else:
        sub_url = "https://www.coupang.com/" + link.attrs["href"]

    response = requests.get(sub_url, headers=headers)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")

    # 브랜드명은 있을 수도 있고, 없을 수도 있음
    # 중고상품의 경우 태그가 달라짐
    # try-except로 예외 처리

    try:
        brand_name = soup.select_one("a.prod-brand-name").text
    except:
        brand_name = ""

    brand_name = soup.select_one("a.prod-brand-name").text

    # 상품명
    product_name = soup.select_one("h2.prod-buy-header__title").text

    # 가격
    product_price = soup.select_one("span.total-price > strong").text

    print(brand_name, product_name, product_price)
"""
import requests
from bs4 import BeautifulSoup
import pyautogui

keyword = pyautogui.prompt("검색어를 입력하세요>>>")

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Accept-Language": "ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3",
}
main_url = "https://www.coupang.com/np/search?component=&q={keyword}&channel=user"

# 헤더에 User-Agent를 추가하지 않으면 오류가 남(멈춰버림)
response = requests.get(main_url, headers=headers)
html = response.text
soup = BeautifulSoup(html, "html.parser")
links = soup.select("a.search-product-link")  # select의 결과는 리스트 자료형
print(links)
for link in links:
    # 광고상품 제거
    if len(link.select("span.ad-badge-text")) > 0:
        print("광고상품입니다.")
    else:
        sub_url = "https://www.coupang.com/" + link.attrs["href"]
        response = requests.get(sub_url, headers=headers)
        html = response.text
        soup = BeautifulSoup(html, "html.parser")

        # 브랜드명은 있을 수도 있고, 없을 수도 있음
        # 중고상품의 경우 태그가 달라짐
        # try-except로 예외 처리

        try:
            brand_name = soup.select_one("a.prod-brand-name").text
        except:
            brand_name = ""

        brand_name = soup.select_one("a.prod-brand-name").text

        # 상품명
        product_name = soup.select_one("h2.prod-buy-header__title").text

        # 가격
        try:
            product_price = soup.select_one("span.total-price > strong").text
        except:
            product_price = 0
        print(brand_name, product_name, product_price)
"""
import requests
from bs4 import BeautifulSoup
import pyautogui

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Accept-Language": "ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3",
}

keyword = pyautogui.prompt("검색어를 입력하세요>>>")

rank = 1
done = False
# 헤더에 User-Agent를 추가하지 않으면 오류가 남(멈춰버림)


for page in range(1, 5):
    if done == True:
        break
    print(page, "번째 페이지입니다.")
    main_url = f"https://www.coupang.com/np/search?&q={keyword}&page={page}"
    response = requests.get(main_url, headers=headers)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    links = soup.select("a.search-product-link")  # select의 결과는 리스트 자료형
    print(links)
    for link in links:
        # 광고상품 제거
        if len(link.select("span.ad-badge-text")) > 0:
            print("광고상품입니다.")
        else:
            sub_url = "https://www.coupang.com/" + link.attrs["href"]
            response = requests.get(sub_url, headers=headers)
            html = response.text
            soup = BeautifulSoup(html, "html.parser")

            # 브랜드명은 있을 수도 있고, 없을 수도 있음
            # 중고상품의 경우 태그가 달라짐
            # try-except로 예외 처리

            try:
                brand_name = soup.select_one("a.prod-brand-name").text
            except:
                brand_name = ""

            brand_name = soup.select_one("a.prod-brand-name").text

            # 상품명
            product_name = soup.select_one("h2.prod-buy-header__title").text

            # 가격
            try:
                product_price = soup.select_one("span.total-price > strong").text
            except:
                product_price = 0
            print(brand_name, product_name, product_price)
            rank = rank + 1
            if rank > 100:
                done = True
                break
"""
